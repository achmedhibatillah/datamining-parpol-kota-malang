# Pengambilan Data Media Sosial
# By. Achmed Hibatillah

import json
import pandas as pd
import requests
import re
from io import BytesIO
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill

# File JSON diletakkan di sini
# JSON untuk per akun dengan format seperti yang terletak dalam direktori /data
# Data dapat diperoleh melalui https://console.apify.com
files = {
    "data/pkb.json": "pkb",
    "data/psi.json": "psi",
    "data/gerindra.json": "gerindra",
    "data/pdip.json": "pdip"
}

excel_file = "analisis_konten.xlsx"
sheets = {}

def is_valid_url(url: str) -> bool:
    return isinstance(url, str) and re.match(r"^https?://", url) is not None

for file, sheet in files.items():
    try:
        with open(file, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"[!] File {file} tidak ditemukan, dilewati...")
        continue

    rows = []
    for post in data:
        hashtags = " ".join([f"#{h}" for h in post.get("hashtags", [])])
        tagged_users = " ".join([f"@{u['username']}" for u in post.get("taggedUsers", [])])
        owner_username = f"@{post.get('ownerUsername','')}"

        row = {
            "Tanggal unggah": post.get("timestamp", ""),
            "Type": post.get("type", ""),
            "Capture": post.get("displayUrl", ""),
            "Caption": post.get("caption", ""),
            "Thumbnail (Kata-kata)": post.get("alt", ""),
            "Subtitle Video": "",
            "dimensionsHeight": post.get("dimensionsHeight", ""),
            "dimensionsWidth": post.get("dimensionsWidth", ""),
            "hashtags": hashtags,
            "locationName": post.get("locationName", ""),
            "ownerUsername": owner_username,
            "taggedUsers": tagged_users,
            "isCommentsDisabled": post.get("isCommentsDisabled", False),
            "isSponsored": post.get("isSponsored", False),
            "Like": post.get("likesCount", 0),
            "Comment": post.get("commentsCount", 0),
            "Share": "",
            "videoDuration": post.get("videoDuration", 0),
            "videoViewCount": post.get("videoViewCount", 0),
            "Lihat..": post.get("url", "")
        }
        rows.append(row)

    rows.sort(key=lambda x: x["Tanggal unggah"], reverse=True)
    sheets[sheet] = pd.DataFrame(rows)

with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
    for sheet, df in sheets.items():
        df.to_excel(writer, sheet_name=sheet, index=False)

wb = load_workbook(excel_file)

for sheet_name in sheets.keys():
    ws = wb[sheet_name]

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        row[14].alignment = Alignment(horizontal="right")
        row[15].alignment = Alignment(horizontal="right")
        row[16].alignment = Alignment(horizontal="right")
        row[17].alignment = Alignment(horizontal="right")
        row[18].alignment = Alignment(horizontal="right")

        owner_cell = row[10]
        if owner_cell.value == f"@{sheet_name}":
            owner_cell.font = Font(color="0000FF")
        else:
            owner_cell.font = Font(color="FF0000")

        lihat_cell = row[19]
        if is_valid_url(lihat_cell.value):
            ws.cell(row=lihat_cell.row, column=20).hyperlink = lihat_cell.value
            ws.cell(row=lihat_cell.row, column=20).value = "Lihat.."
            ws.cell(row=lihat_cell.row, column=20).style = "Hyperlink"

        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        row[8].alignment = Alignment(wrap_text=True, vertical="top")

        for i, cell in enumerate(row):
            if i not in [0,3,4,14,15,16,17,18,19]:
                cell.alignment = Alignment(wrap_text=False)

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        url = row[2].value
        if is_valid_url(url):
            try:
                r = requests.get(url, timeout=10)
                if r.status_code == 200 and r.headers.get("Content-Type", "").startswith("image/"):
                    try:
                        img = Image.open(BytesIO(r.content))
                        img.thumbnail((80, 80))
                        img_path = f"/tmp/tmp_{sheet_name}_{i}.png"
                        img.save(img_path)

                        xl_img = XLImage(img_path)
                        xl_img.anchor = f"C{i}"
                        ws.add_image(xl_img)

                        row[2].value = None
                        ws.row_dimensions[i].height = 70
                        ws.column_dimensions["C"].width = 15
                    except Exception as e:
                        print(f"[!] Gagal proses gambar row {i} di {sheet_name}: {e}")
                        row[2].value = None
                else:
                    print(f"[!] Row {i} di {sheet_name}: URL bukan gambar ({url})")
                    row[2].value = None
            except Exception as e:
                print(f"[!] Gagal ambil gambar row {i} di {sheet_name}: {e}")
                row[2].value = None
        else:
            if url:
                print(f"[!] Row {i} di {sheet_name}: Capture bukan URL ({url})")
            row[2].value = None

    last_row = ws.max_row + 1
    ws[f"N{last_row}"] = "Total"
    ws[f"N{last_row}"].font = Font(bold=True)
    for col, col_letter in zip([14,15,16,17,18], ["O","P","Q","R","S"]):
        ws[f"{col_letter}{last_row}"] = f"=SUM({col_letter}2:{col_letter}{last_row-1})"
        ws[f"{col_letter}{last_row}"].font = Font(bold=True)
        ws[f"{col_letter}{last_row}"].alignment = Alignment(horizontal="right")

for ws in wb.worksheets:
    max_len = max(len(str(cell.value)) for cell in ws['A'])
    ws.column_dimensions['A'].width = max_len + 2
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30

wb.save(excel_file)
print("[v] Data berhasil diekspor ke Excel.")